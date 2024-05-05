from datetime import datetime
from django.utils.html import strip_tags
from django.shortcuts import render, redirect
from main import models
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook


@login_required(login_url='dashboard:log_in')
def index(request):
    quizzes = models.Quiz.objects.filter(author=request.user)
    users = models.User.objects.all().count()
    quiz = models.Quiz.objects.all()
    questions = models.Question.objects.all()
    context = {
        'users':users,
        'quizzes':quizzes,
        'quiz':quiz,
        'questions':questions,
    }

    return render(request, 'index.html', context)

# ---------Quiz----------
def quiz_create(request):
    if request.method == 'POST':
        name = request.POST.get('name') 
        if request.user.is_authenticated:
            quiz = models.Quiz.objects.create(name=name, author=request.user)
            return redirect('quiz_detail', code=quiz.code)
        else:
            return redirect('login')
    return render(request, 'quiz/quiz_create.html')


def quiz_detail(request, code):
    quiz = models.Quiz.objects.get(code=code)
    questions = models.Question.objects.filter(quiz=quiz)
    options = models.Option.objects.filter(question__in=questions)
    context = {
        'quiz':quiz,
        'questions':questions,
        'options':options
    }
    return render(request, 'quiz/quiz_detail.html', context)

def quiz_list(request):
    quizzes = models.Quiz.objects.all()
    return render(request, 'quiz/quiz_list.html', {'quizzes':quizzes})


def quiz_update(request, code):
    quiz = models.Quiz.objects.get(code=code)
    if request.method == 'POST':
        quiz.name = request.POST['name']
        quiz.save()
        return redirect('quiz_detail', quiz.code)


def quiz_delete(request, code):
    quiz = models.Quiz.objects.get(code=code)
    quiz.delete()
    return redirect('quiz_list')


#---------Question----------
def question_create(request, code):
    quiz = models.Quiz.objects.get(code=code)
    if request.method == 'POST':
        question = models.Question.objects.create(
            name=request.POST['question'],
            quiz=quiz
            )
        models.Option.objects.create(
            name = request.POST['correct_option'],
            question = question,
            is_correct = True
        )
        for option in request.POST.getlist('options'):
            models.Option.objects.create(
            name = option,
            question = question,
            is_correct = False
            )
        return redirect('quiz_detail', quiz.code)
    return render(request, 'question/question_create.html', {'quiz':quiz})


def question_detail(request, code):
    question = models.Question.objects.get(code=code)
    return render(request, 'question/question_detail.html', {'question':question})


def question_update(request, code):
    question = models.Question.objects.get(code=code)
    print(request.POST)
    if request.method == 'POST':
        question.name = request.POST.get('name')
        question.save()
        return redirect('quiz_detail', question.code)
    

def question_delete(request, code):
    question = models.Question.objects.get(code=code)
    question.delete()
    return redirect('question_detail', question.code)


#-------Answer--------
def answer_list(request, code):
    answers = models.Answer.objects.filter(quiz__code = code)
    context = {'answers':answers}
    return render(request, 'answer/answer_list.html', context)


def answer_detail(request, code):
    answer = models.Answer.objects.get(code = code)
    details = models.AnswerDetail.objects.filter(answer=answer)
    correct = 0
    in_correct = 0

    for detail in details:
        if detail.is_correct:
            correct += 1
        else:
            in_correct+=1
    correct_percentage = int(correct * 100 / details.count())
    in_correct_percentage = 100 - correct_percentage

    context = {
        'answer':answer,
        'details':details,
        'correct':correct,
        'in_correct':in_correct,
        'total':details.count(),
        'correct_percentage':int(correct_percentage),
        'in_correct_percentage':in_correct_percentage
    }

    return render(request, 'answer/answer_detail.html', context)




def write_excel(request, code):
    answers = models.Answer.objects.filter(quiz__code=code)
    details = models.AnswerDetail.objects.filter(answer__in=answers)
    user_name = []
    email = []
    phone = []
    
    for answer in answers:
        user_name.append(answer.user_name)
        email.append(answer.email) 
        phone.append(answer.phone)

    correct = 0
    in_correct = 0

    for detail in details:
        if detail.is_correct:
            correct += 1
        else:
            in_correct += 1

    quiz_name = strip_tags(answers[0].quiz.name)
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{quiz_name}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active  

    ws['A1'] = 'User_name'
    ws['C1'] = 'Email'
    ws['D1'] = 'Phone'
    ws['E1'] = 'Total questions'
    ws['F1'] = 'Correct'
    ws['G1'] = 'Incorrect'

    for i in range(len(user_name)):
        ws.cell(row=i+2, column=1, value=user_name[i])
        ws.cell(row=i+2, column=2, value=email[i])
        ws.cell(row=i+2, column=3, value=phone[i])
        ws.cell(row=i+2, column=4, value=len(details))
        ws.cell(row=i+2, column=5, value=correct)
        ws.cell(row=i+2, column=6, value=in_correct)

    wb.save(filename)

    return redirect('dashboard:quiz_list')







#---------User----------
def log_in(request):
    if request.method == 'POST':  
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username = username, password = password)
        if user is not None:
            login(request, user)
            return redirect('dashboard:index')
    return render(request, 'auth/login.html')


def log_out(request):
    logout(request)
    return redirect('front:index')
